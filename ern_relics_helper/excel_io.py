from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from .models import Relic, normalize_text, parse_bool


HEADER_ALIASES = {
    "unique_id": {"unique id", "unique_id", "id", "遺物id", "遺物 ID"},
    "retained": {"保留", "保留標記", "已保留", "retained", "keep"},
    "newly_retained": {"新增保留", "newly_retained", "new keep"},
    "relic_type": {"種類", "遺物種類", "type", "relic_type"},
    "color": {"顏色", "color"},
    "mode": {"一般/深夜", "模式", "mode"},
    "total_score": {"總評分", "score", "total_score"},
    "keep_reasons": {"保留原因", "keep_reasons", "reason"},
    "notes": {"註記", "備註", "notes"},
}


def read_relics(path: str | Path, sheet_name: str | None = None) -> list[Relic]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        header_row_index, headers, term_columns = find_relic_header(sheet)
        if not headers and not term_columns:
            raise ValueError("找不到遺物清單標題列，至少需要 unique_id/ID 或 詞條1..詞條6 欄位。")

        relics: list[Relic] = []
        for row_number, row in enumerate(sheet.iter_rows(min_row=header_row_index + 1, values_only=True), start=header_row_index + 1):
            terms = read_terms(row, term_columns, headers.get("terms"))
            if not any(normalize_text(value) for value in row) or not terms:
                continue
            unique_id = read_cell(row, headers.get("unique_id")) or f"row-{row_number}"
            relics.append(
                Relic(
                    unique_id=unique_id,
                    retained=parse_bool(read_cell(row, headers.get("retained"))),
                    relic_type=read_cell(row, headers.get("relic_type")),
                    color=read_cell(row, headers.get("color")),
                    mode=read_cell(row, headers.get("mode")),
                    terms=tuple(terms[:6]),
                )
            )
        return relics
    finally:
        workbook.close()


def write_relics(path: str | Path, relics: list[Relic]) -> None:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "遺物清單"
    sheet.append(
        [
            "unique_id",
            "保留標記",
            "新增保留",
            "遺物種類",
            "顏色",
            "一般/深夜",
            "詞條1",
            "詞條2",
            "詞條3",
            "詞條4",
            "詞條5",
            "詞條6",
            "總評分",
            "保留原因",
            "註記",
        ]
    )
    for relic in relics:
        terms = list(relic.terms)[:6]
        terms.extend([""] * (6 - len(terms)))
        sheet.append(
            [
                relic.unique_id,
                "是" if relic.retained else "",
                "是" if relic.newly_retained else "",
                relic.relic_type,
                relic.color,
                relic.mode,
                *terms,
                relic.total_score,
                "；".join(relic.keep_reasons),
                "；".join(relic.notes),
            ]
        )
    workbook.save(output_path)


def write_relic_template(path: str | Path) -> None:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "遺物清單"
    sheet.append(
        [
            "unique_id",
            "保留標記",
            "遺物種類",
            "顏色",
            "一般/深夜",
            "詞條1",
            "詞條2",
            "詞條3",
            "詞條4",
            "詞條5",
            "詞條6",
        ]
    )
    sheet.append(["relic-001", "", "", "", "", "", "", "", "", "", ""])
    workbook.save(output_path)


def find_relic_header(sheet) -> tuple[int, dict[str, int], list[int]]:
    for row_index, row in enumerate(sheet.iter_rows(max_row=20, values_only=True), start=1):
        headers: dict[str, int] = {}
        term_columns: list[int] = []
        for column_index, value in enumerate(row):
            text = normalize_text(value)
            lowered = text.lower()
            if not text:
                continue
            for field, aliases in HEADER_ALIASES.items():
                if text in aliases or lowered in aliases:
                    headers[field] = column_index
            if text in {"詞條", "辭條", "terms"} or lowered == "terms":
                headers["terms"] = column_index
            if text.startswith(("詞條", "辭條")) and any(char.isdigit() for char in text):
                term_columns.append(column_index)
        if headers or term_columns:
            return row_index, headers, sorted(term_columns)
    return 0, {}, []


def read_terms(row: tuple[object, ...], term_columns: list[int], combined_column: int | None) -> list[str]:
    values = [read_cell(row, column) for column in term_columns]
    if combined_column is not None:
        values.extend(split_terms(read_cell(row, combined_column)))
    return [value for value in values if value]


def split_terms(value: str) -> list[str]:
    parts = [value]
    for separator in ["；", ";", "\n", "、", "|"]:
        parts = [piece for part in parts for piece in part.split(separator)]
    return [part.strip() for part in parts if part.strip()]


def read_cell(row: tuple[object, ...], index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    return normalize_text(row[index])
