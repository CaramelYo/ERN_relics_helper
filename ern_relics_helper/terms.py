from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from .models import TermRule, normalize_stack_state, normalize_text, parse_score, parse_tags


TERM_ALIASES = {"詞條", "辭條", "term", "name"}
CATEGORY_ALIASES = {"類型", "種類", "詞條種類", "辭條種類", "category"}
STACK_ALIASES = {"疊加", "疊加性", "局外疊加性", "局內疊加性", "是否可以疊加", "stack"}
LOGIC_ALIASES = {"邏輯判斷", "邏輯", "標籤", "tags", "logic"}
SCORE_ALIASES = {"評分", "分數", "score"}


def load_term_rules(path: str | Path, sheet_name: str | None = None) -> dict[str, TermRule]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rules: dict[str, TermRule] = {}
    try:
        sheets = [workbook[sheet_name]] if sheet_name else workbook.worksheets
        for sheet in sheets:
            header_row_index, headers = find_header(sheet)
            if not headers or "term" not in headers:
                continue

            for row in sheet.iter_rows(min_row=header_row_index + 1, values_only=True):
                term = normalize_text(row[headers["term"]])
                if not term:
                    continue
                rule = TermRule(
                    term=term,
                    category=read_cell(row, headers.get("category")),
                    stack_state=normalize_stack_state(read_cell(row, headers.get("stack"))),
                    logic_tags=parse_tags(read_cell(row, headers.get("logic"))),
                    score=parse_score(read_cell(row, headers.get("score"))),
                )
                rules[term] = merge_rule(rules.get(term), rule)
        return rules
    finally:
        workbook.close()


def write_term_rules(path: str | Path, rules: dict[str, TermRule]) -> None:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "遺物詞條對照表"
    sheet.append(["詞條", "詞條種類", "疊加", "邏輯判斷", "評分"])
    for rule in sorted(rules.values(), key=lambda item: item.term):
        sheet.append(
            [
                rule.term,
                rule.category,
                rule.stack_state,
                "；".join(rule.logic_tags),
                rule.score,
            ]
        )
    workbook.save(output_path)


def find_header(sheet) -> tuple[int, dict[str, int]]:
    for row_index, row in enumerate(sheet.iter_rows(max_row=20, values_only=True), start=1):
        aliases = {}
        for column_index, value in enumerate(row):
            text = normalize_text(value)
            lowered = text.lower()
            if text in TERM_ALIASES or lowered in TERM_ALIASES:
                aliases["term"] = column_index
            elif text in CATEGORY_ALIASES or lowered in CATEGORY_ALIASES:
                aliases["category"] = column_index
            elif text in STACK_ALIASES or lowered in STACK_ALIASES or "疊加性" in text:
                aliases["stack"] = column_index
            elif text in LOGIC_ALIASES or lowered in LOGIC_ALIASES:
                aliases["logic"] = column_index
            elif text in SCORE_ALIASES or lowered in SCORE_ALIASES:
                aliases["score"] = column_index
        if "term" in aliases:
            return row_index, aliases
    return 0, {}


def read_cell(row: tuple[object, ...], index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    return normalize_text(row[index])


def merge_rule(existing: TermRule | None, new_rule: TermRule) -> TermRule:
    if existing is None:
        return new_rule
    return TermRule(
        term=existing.term,
        category=existing.category or new_rule.category,
        stack_state=existing.stack_state or new_rule.stack_state,
        logic_tags=existing.logic_tags or new_rule.logic_tags,
        score=existing.score if existing.score != 0 else new_rule.score,
    )
